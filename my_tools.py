from serpapi import GoogleSearch
import json
import openpyxl
import xlsxwriter
from datetime import datetime
import os
from dotenv import load_dotenv
import glob
load_dotenv()

def track_price(product):
    params = {
        "engine": "google_shopping",
        "q": product,
        "location": "Hyderabad, Telangana, India",
        "sort_by": "1",
        "num": "5",
        "google_domain": "google.com",
        "hl": "en",
        "gl": "in",
        "api_key": os.getenv("SERPAPI_KEY")
    }
    try:
        search = GoogleSearch(params)
        results = search.get_dict()
        shopping_results = results.get("shopping_results", [])
        cleaned_data = []
        for item in shopping_results:
            cleaned_item = {
                "product_name": item.get("title"),
                "price_raw" : item.get("price"),
                "price_numeric": item.get("extracted_price"),
                "store": item.get("source"),
                "link": item.get("product_link"),
                "reviews": item.get("reviews"),
                "rating": item.get("rating"),
                "condition": item.get("second_hand_condition", "new")
            }
            cleaned_data.append(cleaned_item)
        if not cleaned_data:
            return json.dumps({"error":"no data found"})
        return json.dumps(cleaned_data, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": f"SerpApi failed: {str(e)}"})


def extract_main_file(file_name: str) -> dict[str, list[str]]:
    """
    Extracts the list of product names from the 3rd column of the main Excel file.
    Args:
        file_name: The path to the .xlsx file (e.g., "book.xlsx").
    """
    print(f"\n[TOOL] Extracting products from {file_name}...")
    try:
        # Load the workbook using openpyxl
        main_data = openpyxl.load_workbook(file_name, data_only=True)
        sheet = main_data.active
        product_names = []
        product_prices = ["My Prices"]
        # Iterate through the 3rd column (C)
        for namerow in sheet.iter_rows(min_col=1, max_col=1):
            cell = namerow[0]
            product_name = cell.value
            # Filter out empty cells or headers if necessary
            if product_name and isinstance(product_name, str) and product_name.lower() != "product name":
                product_names.append(product_name)
        
        for namerow in sheet.iter_rows(min_col=2, max_col=2):
            cell = namerow[0]
            product_price = cell.value
            # Filter out empty cells or headers if necessary
            if product_price and isinstance(product_price, int):
                product_prices.append(product_price)
        fin_dict = dict(zip(product_names, product_prices))
        print(f"[TOOL] Found: {fin_dict}")

        return fin_dict

    except Exception as e:
        print(f"[ERROR] Failed to read Excel file: {e}")
        return {"products": [], "error": str(e)}

def save_search(data_json: str) -> dict[str, str]:
    """
    Saves the final calculated product data to a CSV file.
    Args:
        data_json: A JSON string. The agent usually sends a list of objects 
                   like: '[{"Product": "Laptop", "Average Price": 50000}, ...]'
    """
    print(f"\n[TOOL] Saving final report...")
    now = datetime.now()
    current_datetime = now.strftime("%Y%m%d_%H%M%S")
    filename = f"final_market_analysis_{current_datetime}.xlsx"

    try:
        data = json.loads(data_json)
        if isinstance(data, dict):
            data = list(data.values())[0]
        if not isinstance(data, list):
            data = [data]

        if not data:
            print("[TOOL] Warning: No data to save.")
            return {"status": "Error", "message": "Data list was empty."}

        # We dynamically get headers from the keys of the first item (e.g., "Product Name", "Average Price")
        fieldnames = data[0].keys()
        folder = "verdict"
        if not os.path.exists(folder):
            os.mkdir(folder)
        full_filepath = os.path.join(folder, filename)
        workbook = xlsxwriter.Workbook(full_filepath)
        worksheet = workbook.add_worksheet()
        for col_num, field in enumerate(fieldnames):
            worksheet.write(0, col_num, field)

        row_num = 1 
        for row_data in data:
            for col_num, field in enumerate(fieldnames):
                worksheet.write(row_num, col_num, row_data.get(field, ""))
            row_num += 1
        workbook.close()
        print(f"[SUCCESS] File saved as '{filename}' with {len(data)} rows.")
        return {
            "status": "Success", 
            "message": f"Successfully saved {len(data)} products to {filename}"
        }

    except json.JSONDecodeError:
        print(f"[ERROR] JSON Decode Failed. Input was: {data_json[:100]}...")
        return {"status": "Error", "message": "Agent provided invalid JSON."}
    except Exception as e:
        print(f"[ERROR] File Write Failed: {e}")
        return {"status": "Error", "message": str(e)}


def file_to_analyze():

    folder_path = "verdict"
    file_pattern = os.path.join(folder_path, "final_market_analysis_*.xlsx")
    list_of_files = glob.glob(file_pattern)

    if not list_of_files:
        print("[ERROR] No analysis files found")
        exit()
    latest_file = max(list_of_files, key=os.path.getctime)
    print(f"[INFO] Processing latest file: {latest_file}")
    try:
        main_data = openpyxl.load_workbook(latest_file, data_only=True)
        sheet = main_data.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            print("[ERROR] Sheet is empty")
            return {}
        headers = rows[0]
        name_col_index = -1
        possible_keys = ['Product Name', 'Product', 'name', 'title', 'product_name']
        for index, header in enumerate(headers):
            if header and str(header).strip() in possible_keys:
                name_col_index = index
                break
        if name_col_index == -1:
            print(f"[ERROR] Could not find a 'Product Name' column in headers: {headers}")
            return {}
        market_data = {}
        for row in rows[1:]:
            key = row[name_col_index]
            if not key:
                continue
            row_data = {}
            for i, cell_value in enumerate(row):
                if i != name_col_index:
                    header_name = headers[i] if headers[i] else f"cols{i}"
                    row_data[header_name] = cell_value
            market_data[key] = row_data
        return market_data
    except Exception:
        print(f"[Error] Met an Exception {Exception}")

# Archived for sentiment analysis
# extractor = selectorlib.Extractor.from_yaml_file('selectors.yml')
#
# def scrape_gsmarena_reviews(product_name):
#     # GSMArena is much friendlier, but we still use good headers
#     headers = {
#         'authority': 'www.amazon.in',
#         'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#         'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
#         'referer': 'https://www.amazon.in/'
#     }
#
#     search_query = product_name.replace(" ", "+")
#     url = f"https://www.amazon.in/{search_query}"
#     print(f"[SCRAPE TOOL] Searching for {product_name}...")
#     try:
#         r = requests.get(url, headers=headers, timeout=10)
#     except Exception as e:
#         print(f"Network Error: {e}")
#         return None
#
#     if r.status_code != 200:
#         print(f"Blocked or Error. Status Code: {r.status_code}")
#         return None
#
#     # 2. Extract Data
#     data = extractor.extract(r.text, base_url=url)
#     print(data)
#
# # --- TEST IT ---
# # Note: I'm using a REAL URL for the iPhone 15 Pro Max (iPhone 17 isn't out/reviewed yet)
# test_url = "https://www.amazon.com/apple_iphone_15_pro_max-reviews-12548.php"
# from seleniumbase import SB
# from dateutil import parser as dateparser
#
# # def get_reviews_sb(product_name):
# #     """
# #     Uses SeleniumBase with Undetected ChromeDriver (uc=True) 
# #     to bypass Cloudflare and scrape GSMArena.
# #     """
# #     # 1. Initialize with uc=True (Undetected Mode)
# #     # headless=False is safer for bypassing captchas initially. 
# #     # If you need it hidden, try headless=True, but it's slightly more detectable.
# #     with SB(uc=True, test=True, headless=False) as sb: 
# #
# #         print(f"🤖 SB Browser launched. Searching for: {product_name}...")
# #
# #         # --- STEP 1: SEARCH ---
# #         search_query = product_name.replace(" ", "+")
# #         url = f"https://www.amazon.in/{search_query}"
# #
# #         sb.activate_cdp_mode(url) # Special CDP mode often helps with loading
# #         sb.open(url)
# #
# #         # Handle "Verify you are human" checks automatically if they appear
# #         if sb.is_element_visible('iframe[title*="Cloudflare"]'):
# #             print("CAPTCHA detected. Attempting automatic bypass...")
# #             sb.uc_gui_click_captcha() # SB's built-in captcha clicker
# #
# #         # Scenario A: Redirected directly to product
# #         if "res.php3" not in sb.get_current_url():
# #             print(f"Direct redirect to: {sb.get_current_url()}")
# #
# #         # Scenario B: Search Results List
# #         elif sb.is_element_visible("div.makers ul li a"):
# #             print("Clicking top result...")
# #             sb.click("div.makers ul li a")
# #
# #         else:
# #             print("❌ No results found.")
# #             return None
# #
# #         # --- STEP 2: NAVIGATE TO REVIEWS ---
# #         # Convert Specs URL to Reviews URL
# #         current_url = sb.get_current_url()
# #
# #         return current_url
# #
# from bs4 import BeautifulSoup
# def get_ebay_reviews(product_name):
#     session = requests.Session()
#     session.headers.update({
#         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#         'Accept-Language': 'en-US,en;q=0.9',
#     })
#
#     print(f"🔎 Searching eBay for: {product_name}...")
#
#     # --- STEP 1: SEARCH & FIND PRODUCT ---
#     search_url = f"https://www.ebay.com/sch/i.html?_nkw={product_name.replace(' ', '_')}&_sacat=0"
#
#     try:
#         r = session.get(search_url)
#         soup = BeautifulSoup(r.text, 'html.parser')
#
#         # Find the first item that likely has reviews (look for star ratings in search results)
#         # eBay search results usually have a specific class for items
#         items = soup.select('li.s-item')
#
#         review_page_url = None
#
#         for item in items:
#             # Skip "Shop on eBay" header item
#             if "s-item__title--has-tags" in str(item): continue
#
#             link_tag = item.select_one('a.s-item__link')
#             if not link_tag: continue
#
#             item_url = link_tag['href']
#
#             # Optimization: We need to go to the item page to find the "See all reviews" link.
#             # However, some eBay items are just "Listings" without a catalog product.
#             # We will try the first valid organic result.
#             if "ebay.com/itm/" in item_url:
#                 print(f"   Checking item: {item.select_one('.s-item__title').text[:40]}...")
#
#                 # Go to Item Page
#                 r_item = session.get(item_url)
#                 soup_item = BeautifulSoup(r_item.text, 'html.parser')
#
#                 # --- STEP 2: FIND "SEE ALL REVIEWS" LINK ---
#                 # This is the golden ticket. We look for the link that takes us to the /urw/ page.
#                 # It usually text like "See all 43 reviews"
#
#                 # Try Selector A (Modern Layout)
#                 review_link_tag = soup_item.select_one('a[href*="/urw/"]')
#
#                 # Try Selector B (Classic Layout - looking for text)
#                 if not review_link_tag:
#                     review_link_tag = soup_item.find('a', string=lambda t: t and "See all" in t and "reviews" in t)
#
#                 if review_link_tag:
#                     review_page_url = review_link_tag['href']
#                     print(f"✅ Found Dedicated Review URL: {review_page_url}")
#                     break
#
#         if not review_page_url:
#             print("❌ Could not find a product with a dedicated review page (Catalog ID missing).")
#             print("   Tip: Try a generic product name like 'Sony WH-1000XM5' instead of a specific listing.")
#             return None
#
#         # --- STEP 3: SCRAPE REVIEWS ---
#         print("⬇️ Downloading reviews...")
#         r_reviews = session.get(review_page_url)
#
#         # Use SelectorLib to extract data
#         data = extractor.extract(r_reviews.text, base_url=review_page_url)
#
#         # --- STEP 4: CLEAN DATA ---
#         cleaned_reviews = []
#         if data and data.get('reviews'):
#             for r in data['reviews']:
#                 if not r: continue
#
#                 # Clean Rating (e.g., "5 out of 5 stars" -> "5")
#                 if r.get('rating'):
#                     r['rating'] = r['rating'].split(' ')[0]
#
#                 # Clean Date
#                 if r.get('date'):
#                     try:
#                         r['date'] = dateparser.parse(r['date']).strftime('%d %b %Y')
#                     except:
#                         pass
#
#                 cleaned_reviews.append(r)
#
#             data['reviews'] = cleaned_reviews
#             data['count'] = len(cleaned_reviews)
#             print(data)
#         else:
#             print("⚠️ Page loaded, but selectorlib found no reviews. (Selectors might need update)")
#             return None
#
#     except Exception as e:
#         print(f"Error: {e}")
#         return None
#
# get_ebay_reviews("Sony WH-1000XM5")
